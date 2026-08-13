"""Tool for changing the columns of a sheet.

Adds, renames and deletes whole columns. A column is added at the right hand
end rather than pushed in between two others, because openpyxl moves cells
without moving the formulas that read them: anything that shifts columns
sideways leaves every formula after the shift pointing at the wrong cells.
Deleting shifts by its nature, which is why it is refused whenever a formula
would be left behind reading the wrong thing.
"""

from pathlib import Path
from typing import Literal

from langchain_core.tools import tool
from openpyxl.utils import get_column_letter

from excel_agent.config import resolve_workbook
from excel_agent.tracing import traced
from excel_agent.workbook import (
    WRITE_LOCK,
    columns_referenced,
    find_header_row,
    formula_cells,
    header_map,
    is_blank,
    last_data_row,
    load_book,
    location,
    resolve_sheet,
    save,
)


def column_holding(cell, headers: dict[str, int]) -> str:
    """The name of the column a cell sits in, or its letter if it has none."""
    for name, number in headers.items():
        if number == cell.column:
            return f'"{name}"'
    return f"column {get_column_letter(cell.column)}"


@tool
@traced
def modify_column(
    action: Literal["add", "rename", "delete"],
    name: str | None = None,
    new_name: str | None = None,
    workbook: str | None = None,
    sheet: str | None = None,
) -> str:
    """Add, rename or delete a whole column.

    Call inspect_sheet first, so the column names you use are the real ones.
    This changes the columns themselves, not the values in them: use
    modify_row to put values into a column once it exists.

    Args:
        action: What to do. "add" puts a new empty column after the last one,
            "rename" changes a column's name and leaves its data alone,
            "delete" removes a column and everything in it.
        name: The column to act on. For "add" this is the name the new column
            will have. For "rename" and "delete" it is the column that is
            already there, spelled exactly as inspect_sheet reports it.
        new_name: The name a renamed column takes. Only used by "rename".
        workbook: Which workbook to change, by file name. Leave this out to
            change the one being worked on.
        sheet: Which sheet to change, by name. Leave this out to change the
            sheet the workbook opens on.

    Returns:
        A sentence saying what changed, or an explanation of why nothing was
        changed. Nothing is written to the file when the answer is an
        explanation, so you are free to correct the arguments and try again.

        Deleting a column throws its data away for good. A column that some
        formula depends on cannot be deleted at all, and the explanation says
        which formula is in the way.

    Examples:
        modify_column(action="add", name="Notes")
        modify_column(action="rename", name="Units", new_name="Quantity")
        modify_column(action="delete", name="Notes")
    """
    try:
        path = resolve_workbook(workbook)
    except ValueError as explanation:
        return str(explanation)

    with WRITE_LOCK:
        return apply_column_change(action, name, new_name, path, sheet)


def apply_column_change(
    action: str,
    name: str | None,
    new_name: str | None,
    path: Path,
    sheet_name: str | None = None,
) -> str:
    """Do the work of modify_column, with the write lock already held."""
    book = load_book(path)

    try:
        sheet = resolve_sheet(book, sheet_name)
    except ValueError as explanation:
        return str(explanation)

    header_row = find_header_row(sheet)
    headers = header_map(sheet, header_row)
    if not headers:
        return (
            f"No column names were found. Row {header_row} is empty, and no "
            "row near the top of the sheet looks like a header."
        )

    if is_blank(name):
        return f"The {action} action needs the name of a column."

    assert name is not None
    name = name.strip()

    if action in ("rename", "delete") and name not in headers:
        return (
            f'There is no column called "{name}". '
            f"The sheet has: {', '.join(headers)}."
        )

    if action == "add":
        if name in headers:
            return (
                f'There is already a column called "{name}". '
                "Two columns of one name cannot be told apart."
            )

        column = max(headers.values()) + 1
        sheet.cell(row=header_row, column=column).value = name
        save(book, path)
        return (
            f'Added a column called "{name}", at {get_column_letter(column)}. '
            "It is empty: use modify_row to put values into it."
        ) + location(sheet, path)

    if action == "rename":
        if is_blank(new_name):
            return "The rename action needs a new_name to give the column."

        assert new_name is not None
        new_name = new_name.strip()

        if new_name in headers and new_name != name:
            return (
                f'There is already a column called "{new_name}". '
                "Two columns of one name cannot be told apart."
            )

        sheet.cell(row=header_row, column=headers[name]).value = new_name
        save(book, path)
        return (
            f'Renamed the column "{name}" to "{new_name}". '
            "Its data has not moved, and the values in it are unchanged."
        ) + location(sheet, path)

    if action == "delete":
        target = headers[name]
        last_row = last_data_row(sheet, header_row)

        # A formula in the column being deleted goes with it, so only the ones
        # that will still be there afterwards matter.
        for cell in formula_cells(sheet, header_row, last_row):
            if cell.column == target:
                continue
            if any(read >= target for read in columns_referenced(str(cell.value))):
                return (
                    f'"{name}" cannot be deleted. {column_holding(cell, headers)} '
                    f"is worked out by a formula that reads it, or a column to "
                    f"the right of it ({cell.coordinate}: {cell.value}). "
                    "Deleting a column slides everything after it one place to "
                    "the left without moving the formulas as well, so that "
                    "formula would end up reading the wrong cells."
                )

        sheet.delete_cols(target)
        save(book, path)
        return (
            f'Deleted the column "{name}" and everything in it, which cannot '
            "be brought back. Every column that was to its right has moved one "
            "place left."
        ) + location(sheet, path)

    return f'Unknown action "{action}". Use add, rename or delete.'


def main() -> None:
    """Try the tool by hand with `python -m excel_agent.tools.columns`.

    Refuses one deletion, then adds a column, renames it and deletes it again,
    so the sheet is left as it was found. Close the file in Excel first,
    otherwise saving fails because Excel holds a lock on it while it is open.
    """
    print("--- a column a formula depends on, so nothing is written ---")
    print(modify_column.invoke({"action": "delete", "name": "Units"}))

    print("\n--- add a column ---")
    print(modify_column.invoke({"action": "add", "name": "Test Column"}))

    print("\n--- rename it ---")
    print(
        modify_column.invoke(
            {"action": "rename", "name": "Test Column", "new_name": "Renamed Column"}
        )
    )

    print("\n--- delete it again ---")
    print(modify_column.invoke({"action": "delete", "name": "Renamed Column"}))


if __name__ == "__main__":
    main()