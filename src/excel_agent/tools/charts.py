"""Tool for putting a chart on a sheet.

A chart holds a range of cells rather than a copy of their values, so Excel
works out what to draw when the file is opened. That is why a calculated
column can be charted even though its results are not stored in the file: the
chart points at the formulas, and Excel evaluates them.

The range is fixed when the chart is made. Rows added afterwards fall outside
it, and the chart has to be made again to take them in.
"""

import re
from pathlib import Path
from typing import Literal

from langchain_core.tools import tool
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from excel_agent.config import resolve_workbook
from excel_agent.workbook import (
    WRITE_LOCK,
    find_header_row,
    header_map,
    is_blank,
    last_data_row,
    load_book,
    load_values,
    location,
    resolve_sheet,
    save,
)

KINDS = {"bar": BarChart, "line": LineChart, "pie": PieChart}

# How far down the next chart goes, in rows. A chart is 7.5cm tall by default,
# which is about fifteen rows, so sixteen clears it.
CHART_DEPTH = 16

ANCHOR = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*$")


def free_anchor(worksheet, headers: dict[str, int], header_row: int) -> str:
    """Where to put a chart so it does not cover the data or another chart.

    Two columns clear of the last named column, then one chart's depth further
    down for each chart already on the sheet. Charts loaded from the file are
    counted too, so a second chart drawn in a later session still lands below
    the first rather than on top of it.
    """
    column = get_column_letter(max(headers.values()) + 2)
    return f"{column}{header_row + len(worksheet._charts) * CHART_DEPTH}"


def is_number(value) -> bool:
    """Whether a value is a number worth plotting."""
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def can_be_plotted(values_sheet, formulas_sheet, column: int, rows) -> bool:
    """Whether a column holds numbers, or works them out.

    A calculated column is plottable even when nothing can be read from it.
    Its results are not in the file until Excel has opened it, but a chart
    points at the cells rather than at their values, so Excel draws it from
    the formulas themselves.
    """
    filled = [
        values_sheet.cell(row=row, column=column).value
        for row in rows
        if not is_blank(values_sheet.cell(row=row, column=column).value)
    ]
    if filled:
        return all(is_number(value) for value in filled)

    return any(
        str(formulas_sheet.cell(row=row, column=column).value or "").startswith("=")
        for row in rows
    )


@tool
def modify_chart(
    action: Literal["add", "remove"],
    values: str | None = None,
    categories: str | None = None,
    kind: Literal["bar", "line", "pie"] = "bar",
    title: str | None = None,
    anchor: str | None = None,
    workbook: str | None = None,
    sheet: str | None = None,
) -> str:
    """Draw a chart on the sheet, or take the charts off it.

    Call inspect_sheet first, so the column names you use are the real ones.
    The chart is placed clear of the data, and reads whole columns: it covers
    every row of data the sheet has when you make it.

    Args:
        action: "add" draws a chart, "remove" takes every chart off the sheet.
        values: The column to plot, by name. It must hold numbers, or be a
            column the sheet works out for itself.
        categories: The column whose values label the bars, points or slices.
            Leave it out to label them 1, 2, 3 and so on.
        kind: "bar", "line" or "pie".
        title: What to call the chart. Left out, it is named after the columns.
        anchor: The cell to put the chart's top left corner in, such as "H2".
            Left out, it goes clear of the data, and below any chart already
            on the sheet rather than on top of it.
        workbook: Which workbook to change, by file name. Leave this out to
            change the one being worked on.
        sheet: Which sheet to change, by name. Leave this out to change the
            sheet the workbook opens on.

    Returns:
        A sentence saying what was drawn, or an explanation of why nothing was
        drawn. Nothing is written to the file when the answer is an
        explanation.

        A chart keeps the range it was given. Rows added later are not in it,
        so draw it again once the data is complete.

    Examples:
        modify_chart(action="add", values="Units", categories="Product")
        modify_chart(action="add", values="Total", categories="Region", kind="pie")
        modify_chart(action="add", values="Units", anchor="H2")
        modify_chart(action="remove")
    """
    try:
        path = resolve_workbook(workbook)
    except ValueError as explanation:
        return str(explanation)

    with WRITE_LOCK:
        return apply_chart_change(
            action, values, categories, kind, title, anchor, path, sheet
        )


def apply_chart_change(
    action: str,
    values: str | None,
    categories: str | None,
    kind: str,
    title: str | None,
    anchor: str | None,
    path: Path,
    sheet_name: str | None = None,
) -> str:
    """Do the work of modify_chart, with the write lock already held."""
    book = load_book(path)

    try:
        worksheet = resolve_sheet(book, sheet_name)
    except ValueError as explanation:
        return str(explanation)

    if action == "remove":
        drawn = len(worksheet._charts)
        if not drawn:
            return f"There are no charts on {worksheet.title} to remove."

        worksheet._charts.clear()
        save(book, path)
        return (
            f"Removed {drawn} chart{'s' if drawn > 1 else ''}. "
            "The data itself is untouched."
        ) + location(worksheet, path)

    if action != "add":
        return f'Unknown action "{action}". Use add or remove.'

    header_row = find_header_row(worksheet)
    headers = header_map(worksheet, header_row)
    if not headers:
        return (
            f"No column names were found. Row {header_row} is empty, and no "
            "row near the top of the sheet looks like a header."
        )

    if is_blank(values):
        return "The add action needs the name of a column to plot, in values."

    wanted = [name for name in (values, categories) if name]
    unknown = [name for name in wanted if name not in headers]
    if unknown:
        return (
            f"Unknown column(s): {', '.join(unknown)}. "
            f"The sheet has: {', '.join(headers)}."
        )

    if kind not in KINDS:
        return f'Unknown kind "{kind}". Use one of: {", ".join(KINDS)}.'

    if anchor and not ANCHOR.match(anchor.strip()):
        return (
            f'"{anchor}" is not a cell. Give one cell for the chart to start '
            'at, such as "H2".'
        )

    last_row = last_data_row(worksheet, header_row)
    if last_row <= header_row:
        return "There are no rows of data to draw a chart from yet."

    rows = range(header_row + 1, last_row + 1)
    assert values is not None
    if not can_be_plotted(load_values(path)[worksheet.title], worksheet, headers[values], rows):
        return (
            f'"{values}" cannot be plotted, because it does not hold numbers. '
            "Choose a column of numbers, or one the sheet works out for itself."
        )

    chart = KINDS[kind]()
    chart.title = title or (
        f"{values} by {categories}" if categories else values
    )
    chart.add_data(
        Reference(
            worksheet,
            min_col=headers[values],
            min_row=header_row,
            max_row=last_row,
        ),
        titles_from_data=True,
    )
    if categories:
        chart.set_categories(
            Reference(
                worksheet,
                min_col=headers[categories],
                min_row=header_row + 1,
                max_row=last_row,
            )
        )

    where = anchor.strip() if anchor else free_anchor(worksheet, headers, header_row)
    worksheet.add_chart(chart, where.upper())
    save(book, path)

    labelled = f", labelled by {categories}" if categories else ""
    return (
        f"Drew a {kind} chart of {values}{labelled}, at {where.upper()}. "
        f"It covers rows {header_row + 1} to {last_row}: rows added after this "
        "are not in it, so draw it again if the data grows."
    ) + location(worksheet, path)


def main() -> None:
    """Try the tool by hand with `python -m excel_agent.tools.charts`.

    Draws a chart, then takes it off again, so the sheet is left as it was
    found. Close the file in Excel first, otherwise saving fails.
    """
    print("--- a column that cannot be plotted, so nothing is drawn ---")
    print(modify_chart.invoke({"action": "add", "values": "Product"}))

    print("\n--- draw a bar chart ---")
    print(
        modify_chart.invoke(
            {"action": "add", "values": "Units", "categories": "Product"}
        )
    )

    print("\n--- take it off again ---")
    print(modify_chart.invoke({"action": "remove"}))


if __name__ == "__main__":
    main()