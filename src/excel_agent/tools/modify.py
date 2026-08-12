"""Tool for changing the sheet.

Adds, edits and removes rows. Every write goes through workbook.save(), so a
backup is taken before the file is touched.
"""

from typing import Literal

from pathlib import Path

from langchain_core.tools import tool
from openpyxl.utils import get_column_letter

from excel_agent.config import resolve_workbook
from excel_agent.tracing import traced
from excel_agent.workbook import (
    WRITE_LOCK,
    copy_row_formulas,
    find_header_row,
    formula_columns,
    header_map,
    last_data_row,
    load_book,
    location,
    resolve_sheet,
    save,
)


def describe(values: dict) -> str:
    """Render the changed columns for the confirmation message."""
    return ", ".join(
        f"{name} = {'(blank)' if value is None else value}"
        for name, value in values.items()
    )


def holds_a_formula(cell) -> bool:
    """Whether a cell works its value out rather than holding one."""
    return isinstance(cell.value, str) and cell.value.startswith("=")


def formula_refusal(names: list[str]) -> str:
    """The standard explanation for declining to write over a calculation."""
    return (
        f"{', '.join(names)} is worked out by a formula in the "
        "sheet, so it cannot be set directly. Change the columns it "
        "is calculated from instead, and leave it out of values."
    )


def column_names(columns: list[int], headers: dict[str, int]) -> list[str]:
    """Turn column numbers back into column names for the confirmation message.

    Falls back to the column letter for a column with no name in the header
    row, which can happen when a calculated column sits outside the table.
    """
    by_number = {number: name for name, number in headers.items()}
    return [by_number.get(number, get_column_letter(number)) for number in columns]


@tool
@traced
def modify_sheet(
    action: Literal["add", "edit", "remove"],
    row: int | None = None,
    values: dict[str, str | int | float | None] | None = None,
    workbook: str | None = None,
    sheet: str | None = None,
) -> str:
    """Add, edit or remove a row in the sheet.

    Call inspect_sheet first, so the row numbers you use are real ones.

    Args:
        action: What to do. "add" puts a new row at the bottom, "edit" changes
            cells in a row that already exists, "remove" deletes a whole row.
        row: The Excel row number to change. Needed for edit and remove, and
            ignored for add.
        values: Column name mapped to new value. Only the columns listed here are
            changed, and columns you leave out keep their current value. Pass
            null as a value to clear a cell. Ignored for remove.
            A cell the sheet works out for itself cannot be set, whether it is
            a whole calculated column or one formula partway down a column.
            inspect_sheet shows such a cell as its formula. Change the columns
            the formula reads from instead, and leave it out of values.
        workbook: Which workbook to change, by file name. Leave this out to
            change the one being worked on, which is what you normally want.
            Name a workbook only when the user named a file, and read it with
            inspect_sheet first: row numbers from one workbook mean nothing in
            another.
        sheet: Which sheet to change, by name. Leave this out to change the
            sheet the workbook opens on. The same warning applies: a row
            number read from one sheet means nothing in another, so read the
            sheet you are about to change.

    Returns:
        A sentence saying what changed, or an explanation of why nothing was
        changed. Nothing is written to the file when the answer is an
        explanation, so you are free to correct the arguments and try again.

    Examples:
        modify_sheet(action="edit", row=5, values={"Units": 20})
        modify_sheet(action="edit", row=5, values={"Notes": null})
        modify_sheet(action="add", values={"Product": "Webcam", "Region": "EU"})
        modify_sheet(action="remove", row=5)
    """

  
    try:
        path = resolve_workbook(workbook)
    except ValueError as explanation:
        return str(explanation)

    with WRITE_LOCK:
        return apply_change(action, row, values, path, sheet)


def apply_change(
    action: str,
    row: int | None,
    values: dict[str, str | int | float | None] | None,
    path: Path,
    sheet_name: str | None = None,
) -> str:
    """Do the work of modify_sheet, with the write lock already held."""
    book = load_book(path)

    try:
        sheet = resolve_sheet(book, sheet_name)
    except ValueError as explanation:
        return str(explanation)

    header_row = find_header_row(sheet)
    headers = header_map(sheet, header_row)
    if not headers:
        # the modification process depends on the column names being decalred
        return (
            f"No column names were found. Row {header_row} is empty, and no "
            "row near the top of the sheet looks like a header."
        )

    last_row = last_data_row(sheet, header_row)

    # Check everything before writing anything, so a rejected call leaves the
    # file exactly as it was.
    if action in ("edit", "remove"):
        if row is None:
            return (
                f"The {action} action needs a row number. "
                "Call inspect_sheet to find the right one."
            )
        if row <= header_row or row > last_row:
            return (
                f"Row {row} does not exist. The sheet has rows "
                f"{header_row + 1} to {last_row}."
            )

    if action in ("add", "edit"):
        if not values:
            return f"The {action} action needs at least one column in values."
        unknown = [name for name in values if name not in headers]
        if unknown:
            return (
                f"Unknown column(s): {', '.join(unknown)}. "
                f"The sheet has: {', '.join(headers)}."
            )

    # A calculated column is protected differently depending on the action,
    # because the two are about to write to different places.
    if action == "add":
        assert values is not None
        # The new row copies its formulas from the last row, so the question
        # is which columns that row calculates.
        calculated = formula_columns(sheet, header_row, last_row)
        blocked = [name for name in values if headers[name] in calculated]
        if blocked:
            return formula_refusal(blocked)

    if action == "edit":
        assert values is not None and row is not None
        # One row is being written to, so the cells of that row are what to
        # look at. Asking the column instead would miss a formula partway down
        # a column whose last row holds a number someone typed over it, and
        # the formula would be replaced without a word about it.
        blocked = [
            name
            for name in values
            if holds_a_formula(sheet.cell(row=row, column=headers[name]))
        ]
        if blocked:
            return formula_refusal(blocked)

    if action == "add":
        assert values is not None
        new_row = last_row + 1

        for name, value in values.items():
            sheet.cell(row=new_row, column=headers[name], value=value)

        # Carry any calculated column down into the new row, skipping the
        # columns that were given a value so a formula cannot overwrite one.
        copied = []
        if last_row > header_row:
            copied = copy_row_formulas(
                sheet,
                source_row=last_row,
                target_row=new_row,
                skip={headers[name] for name in values},
            )

        save(book, path)

        message = f"Added row {new_row} with {describe(values)}."
        if copied:
            names = ", ".join(column_names(copied, headers))
            message += f" Copied the formula in {names} down from row {last_row}."
        message += " Any other column was left blank."
        return message + location(sheet, path)

    if action == "edit":
        assert values is not None and row is not None
        for name, value in values.items():
            # Assigned rather than handed to cell(value=...), which reads None
            # as "no value was given" and leaves the cell as it was. Clearing a
            # cell is something this tool offers, so it has to mean it.
            sheet.cell(row=row, column=headers[name]).value = value
        save(book, path)
        return f"Updated row {row}: {describe(values)}." + location(sheet, path)

    if action == "remove":
        assert row is not None
        sheet.delete_rows(row)
        save(book, path)
        return (
            f"Removed row {row}. The rows below it have shifted up by one, so "
            "any row numbers you read earlier are now out of date. Call "
            "inspect_sheet again before changing anything else by row number."
        ) + location(sheet, path)

    return f'Unknown action "{action}". Use add, edit or remove.'


def main() -> None:
    """Try the tool by hand with `python -m excel_agent.tools.modify`.

    Adds a row, edits it, then removes it again, so the sheet is left as it
    was found. Close the file in Excel first, otherwise saving fails because
    Excel holds a lock on it while it is open.
    """
    # Imported here rather than at the top, so the tool itself does not depend
    # on the other tool.
    from excel_agent.tools.inspect import inspect_sheet

    print("--- a row that does not exist, so nothing is written ---")
    print(modify_sheet.invoke({"action": "edit", "row": 9999, "values": {"Region": "EU"}}))

    print("\n--- a column that does not exist, so nothing is written ---")
    print(modify_sheet.invoke({"action": "add", "values": {"Profit": 10}}))

    print("\n--- add a row ---")
    print(
        modify_sheet.invoke(
            {
                "action": "add",
                "values": {
                    "Product": "Test Row",
                    "Region": "EU",
                    "Units": 3,
                    "Unit Price": 10,
                },
            }
        )
    )

    # The same workbook the tool calls above reached for, since none of them
    # names one.
    sheet = resolve_sheet(load_book(resolve_workbook()))
    assert sheet is not None
    added_row = last_data_row(sheet, find_header_row(sheet))

    print(f"\n--- edit row {added_row} ---")
    print(modify_sheet.invoke({"action": "edit", "row": added_row, "values": {"Units": 8}}))

    print("\n--- the sheet now, last few rows ---")
    print(inspect_sheet.invoke({"start_row": added_row - 2}))

    print(f"\n--- remove row {added_row} again ---")
    print(modify_sheet.invoke({"action": "remove", "row": added_row}))


if __name__ == "__main__":
    main()