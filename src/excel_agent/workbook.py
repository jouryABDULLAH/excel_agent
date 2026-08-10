"""Enables Excel file access for the agent.

This module holds the rules for opening and saving the workbook so that the
tools do not have to repeat them.
"""

import os
import re
import threading
from datetime import datetime
from pathlib import Path
from shutil import copyfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string

from excel_agent.config import BACKUP_DIR, BACKUP_KEEP

WRITE_LOCK = threading.RLock()

# Used when no row near the top of the sheet looks like a header.
DEFAULT_HEADER_ROW = 1


HEADER_SEARCH_DEPTH = 10

# What a backup is called: the workbook's name, when it was taken, and the
# real suffix, so a backup opens by double clicking like any other workbook.
STAMP_FORMAT = "%Y%m%d-%H%M%S"
STAMP_PATTERN = re.compile(r"-\d{8}-\d{6}")

# An A1 style cell reference: up to three letters, a row number, and the
# dollar signs that pin either of them. The lookbehind keeps it from biting a
# piece out of a longer word.
CELL_REFERENCE = re.compile(r"(?<![A-Za-z0-9_$])\$?([A-Z]{1,3})\$?(\d+)")

# The workbooks backed up so far this session, so each one is copied once
# however many times it is written to.
_backed_up: set[Path] = set()


def load_values(path: Path):
    """Open the workbook with formula results in place of the formulas.

    Used for reading values. The returned workbook is tagged as read only,
    because saving it would write the cached values over the formulas.
    """
    book = load_workbook(path, data_only=True)

    setattr(book, "_agent_read_only", True)
    return book


def load_book(path: Path):
    """Open the workbook for editing, with the formulas kept as formulas."""
    return load_workbook(path)


def resolve_sheet(book, name: str | None = None):
    """Pick a sheet out of an open workbook by name.

    Returns the active sheet the workbook opens on when given nothing. Names are matched ignoring
    case and surrounding spaces, the same way a workbook name is.

    Raises ValueError, with a message worth showing to the model, when the
    name reaches no sheet.
    """
    
    if not name or not name.strip():
        sheet = book.active
        if sheet is None:
            raise ValueError("This workbook has no sheets.")
        return sheet

    wanted = name.strip()
    for title in book.sheetnames:
        if title.lower() == wanted.lower():
            return book[title]

    available = ", ".join(book.sheetnames) or "no sheets at all"
    raise ValueError(
        f'There is no sheet called "{name}". The workbook has: {available}.'
    )


def location(worksheet, path) -> str:
    """Where a change was written, for the end of a confirmation.

    A tool that is told which workbook to change says so back. Leaving the
    argument out is easy, and a write that landed in the wrong file would
    otherwise be silent.
    """
    return f" ({worksheet.title} in {path.name})"


def is_blank(value) -> bool:
    """Whether a cell holds nothing worth reading.

    A cell that was never filled in reads as None, but one emptied by Excel or
    by another program can come back as an empty string instead, and a cell
    holding nothing but spaces looks empty to anyone reading the sheet. All
    three are the same thing, and treating them differently would mean a row
    cleared one way behaved differently from a row cleared the other, while
    looking identical in the table the model is shown.

    A zero is not blank. Neither is False. Both are values someone meant.
    """
    if value is None:
        return True
    return isinstance(value, str) and not value.strip()


def find_header_row(sheet, search_depth: int = HEADER_SEARCH_DEPTH) -> int:
    """Find the row that holds the column names.

    Walks down from the top and takes the first row with at least two filled
    cells, all of them text, and something filled in the row below it. A
    title in A1 fails the two cell test and blank rows fail it as well, so
    sheets that do not start with their header are still read correctly.

    Falls back to DEFAULT_HEADER_ROW when no row near the top looks like one.
    """
    for row in range(1, min(search_depth, sheet.max_row) + 1):
        filled = [cell.value for cell in sheet[row] if not is_blank(cell.value)]

        if len(filled) < 2:
            continue
        if not all(isinstance(value, str) for value in filled):
            continue
        if row >= sheet.max_row:
            continue
        if all(is_blank(cell.value) for cell in sheet[row + 1]):
            continue

        return row

    return DEFAULT_HEADER_ROW


def header_map(sheet, header_row: int) -> dict[str, int]:
    """Map each column name in the header row to its Excel column number.

    Looking columns up by name means no column letters are hardcoded, so a
    column moving in the file does not break the tools. Names are stripped,
    because a header typed as "Total " would otherwise only match when the
    trailing space is included.

    A cell with nothing in it names no column, so it is left out. Keeping it
    would put a column called "" in the map, which nothing could ask for by
    name and which would sit in the middle of every table read from the sheet.
    """
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[header_row]
        if not is_blank(cell.value)
    }


def last_data_row(sheet, header_row: int) -> int:
    """Find the last row that actually holds data.

    sheet.max_row counts rows that carry only formatting, so a sheet styled
    well past its contents reports far more rows than it has. Walking back up
    from there finds the last row with something in it, and returns the
    header row itself when the sheet holds no data at all.

    A row someone emptied is not the end of the data, whichever way it was
    emptied. That means a row cleared to empty strings is passed over exactly
    like one cleared to nothing at all, and the next row added takes its place
    rather than landing underneath it.
    """
    for row in range(sheet.max_row, header_row, -1):
        if any(not is_blank(cell.value) for cell in sheet[row]):
            return row

    return header_row


def formula_columns(sheet, header_row: int, last_row: int) -> set[int]:
    """Column numbers whose values are worked out by a formula.

    Read from the last row of data, on the basis that a calculated column is
    calculated the whole way down. Used to stop a caller writing a number
    into a cell the sheet works out for itself.
    """
    if last_row <= header_row:
        return set()

    return {
        cell.column
        for cell in sheet[last_row]
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def formula_cells(sheet, header_row: int, last_row: int):
    """Every cell in the used range that works its value out.

    formula_columns only reads the last row, which is enough to spot a column
    that is calculated all the way down. Deleting a column needs more than
    that: one formula anywhere in the sheet is enough to make the deletion
    wrong, so every row has to be looked at.
    """
    for row in range(header_row, last_row + 1):
        for cell in sheet[row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                yield cell


def columns_referenced(formula: str) -> set[int]:
    """The column numbers a formula reads from.

    Matches A1 style references, with or without the dollar signs that pin
    them, and both ends of a range. Two things it reads too greedily: a
    function whose name ends in digits, such as LOG10, and a reference to
    another sheet, such as Notes!A1, are both taken for references to this
    sheet. Both make a caller refuse a deletion it could have allowed, which
    is the safe way round to be wrong.
    """
    return {
        column_index_from_string(letters)
        for letters, _ in CELL_REFERENCE.findall(formula)
    }


def copy_row_formulas(
    sheet,
    source_row: int,
    target_row: int,
    skip: set[int] | None = None,
) -> list[int]:
    """Copy every formula in one row down into another row.

    Used when adding a row, so a new row picks up whatever calculated columns
    the sheet already has. Cell references are shifted to match the new row,
    so a formula reading =D11*E11 becomes =D12*E12.

    Columns listed in skip are left alone, which is how a value the caller
    asked for avoids being overwritten by a formula.

    Returns the column numbers that received a formula.
    """
    skip = skip or set()
    copied = []

    for cell in sheet[source_row]:
        if cell.column in skip:
            continue
        if not isinstance(cell.value, str) or not cell.value.startswith("="):
            continue

        target = sheet.cell(row=target_row, column=cell.column)
        target.value = Translator(
            cell.value, origin=cell.coordinate
        ).translate_formula(target.coordinate)
        copied.append(cell.column)

    return copied


def backups_of(path: Path) -> list[Path]:
    """Every backup taken of one workbook, oldest first.

    Found by name: the workbook's own name, then a stamp of the date and time.
    The stamp is matched exactly rather than with a wildcard, so the backups
    of "sales.xlsx" are not confused with those of "sales-20260101-090000.xlsx"
    and deleted in its place when the older ones are cleared out.
    """

    # no backups taken yet
    if not BACKUP_DIR.is_dir():
        return []

    found = [
        backup
        for backup in BACKUP_DIR.glob(f"{path.stem}-*{path.suffix}")
        if STAMP_PATTERN.fullmatch(backup.stem[len(path.stem) :])
    ]
    # The stamp reads year first, so sorting by name sorts by age.
    return sorted(found)


def backup_once(path: Path) -> Path | None:
    """Copy a workbook into the backups folder the first time it is written to.

    Returns the backup path, or None if this workbook was already backed up
    this session. Backing up once means the copy is of the file as it was
    found, rather than of a change made a moment ago.

    The oldest backups of that workbook are deleted once there are more than
    config.BACKUP_KEEP of them, so the folder does not grow forever.
    """
    with WRITE_LOCK:
        path = path.resolve()
        if path in _backed_up:
            return None

        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime(STAMP_FORMAT)
        backup_path = BACKUP_DIR / f"{path.stem}-{stamp}{path.suffix}"
        copyfile(path, backup_path)
        _backed_up.add(path)

        for old in backups_of(path)[:-BACKUP_KEEP]:
            old.unlink()

        return backup_path


def save(book, path: Path) -> None:
    """Save an editable workbook, taking a backup first.

    Raises ValueError if given a workbook came from load_values, since
    saving that one would destroy every formula in the file.

    Writes to a temporary file and then renames it over the real one. The
    rename is a single step as far as the operating system is concerned, so
    anything reading the file sees either the old version or the new one,
    never half of each.
    """
    if getattr(book, "_agent_read_only", False):
        raise ValueError(
            "This workbook was opened for reading only. "
            "Use load_book() to make changes."
        )

    with WRITE_LOCK:
        backup_once(path)

        being_written = path.with_suffix(path.suffix + ".writing")
        book.save(being_written)
        os.replace(being_written, path)