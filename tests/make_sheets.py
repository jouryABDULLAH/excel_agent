"""Builds the workbooks the tests run against.

Each function writes one variant of a sheet into a directory the caller owns,
normally pytest's tmp_path, and returns the path it wrote. Nothing here reads
or writes data/sample.xlsx, and none of these files are committed: a test that
wants a sheet builds it.

"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill


def _save(book: Workbook, directory: Path, name: str) -> Path:
    """Write a workbook into the caller's directory and return its path."""
    path = Path(directory) / name
    book.save(path)
    return path


def clean_table(directory: Path) -> Path:
    """A header in row 1 with data straight under it.

    The shape every other variant is a departure from. Header row 1, data in
    rows 2 to 6, nothing else in the file.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Sales"

    sheet.append(["ID", "Product", "Region", "Units", "Unit Price"])
    sheet.append([1001, "Laptop Stand", "EU", 12, 24.5])
    sheet.append([1002, "USB-C Hub", "EU", 30, 18.0])
    sheet.append([1003, "Monitor Arm", "US", 7, 89.99])
    sheet.append([1004, "Keyboard", "APAC", 40, 55.0])
    sheet.append([1005, "Webcam", "US", 18, 42.0])

    return _save(book, directory, "clean_table.xlsx")


def title_above_header(directory: Path) -> Path:
    """A title in A1 and a blank row, with the real header in row 3.

    The case the two cell minimum was written for: a title fills one cell, so
    it is skipped, and the header below it is found. Header row 3, data in
    rows 4 to 7.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Sales"

    sheet["A1"] = "Q1 Sales Report"
    sheet.append([])
    sheet.append(["ID", "Product", "Region", "Units"])
    sheet.append([1001, "Laptop Stand", "EU", 12])
    sheet.append([1002, "USB-C Hub", "EU", 30])
    sheet.append([1003, "Monitor Arm", "US", 7])
    sheet.append([1004, "Keyboard", "APAC", 40])

    return _save(book, directory, "title_above_header.xlsx")


def numeric_headers(directory: Path) -> Path:
    """A header whose column names are years, so they are numbers not text.

    Detection reads a header as a row of text, and 2024 is not text, so row 2
    is passed over and so is every row of data below it. The fallback then
    lands on the title in row 1, and the title becomes the only column name.
    A sheet read this way is read wrongly, which is the point of the fixture.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Revenue"

    sheet["A1"] = "Revenue by year"
    sheet.append(["Product", 2024, 2025])
    sheet.append(["Laptop Stand", 12000, 15000])
    sheet.append(["USB-C Hub", 8000, 9500])
    sheet.append(["Monitor Arm", 21000, 19000])

    return _save(book, directory, "numeric_headers.xlsx")


def one_column(directory: Path) -> Path:
    """A single column of data, under a title.

    A header row of one cell cannot pass a test that asks for two, so no row
    here is ever taken for a header and the fallback lands on the title. The
    title is then treated as the column name and the real header, "Product",
    is treated as data. Another sheet the tools read wrongly today.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Inventory"

    sheet.append(["Inventory"])
    sheet.append(["Product"])
    sheet.append(["Laptop Stand"])
    sheet.append(["USB-C Hub"])
    sheet.append(["Monitor Arm"])

    return _save(book, directory, "one_column.xlsx")


def headerless_grid(directory: Path) -> Path:
    """Data with no header row at all.

    Nothing in the sheet names a column, so the fallback picks row 1 and the
    first row of data is read as the column names. It is a silent misread:
    every tool answers as though the sheet had a header, and the first row of
    real data disappears from the results. need to add grid tools for this to work.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Grid"

    sheet.append(["Laptop Stand", 12, 24.5])
    sheet.append(["USB-C Hub", 30, 18.0])
    sheet.append(["Monitor Arm", 7, 89.99])
    sheet.append(["Keyboard", 40, 55.0])

    return _save(book, directory, "headerless_grid.xlsx")


def formulas_last_row_overwritten(directory: Path) -> Path:
    """A calculated column whose last row was typed over by hand.

    Total is a formula in rows 2 to 5 and a plain number in row 6. Because a
    calculated column is recognised by reading the last row of data, this
    sheet reports no calculated columns at all, and the formulas in the middle
    of the column are left unguarded. This is the sheet the per-cell check in
    edit exists for.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Sales"

    sheet.append(["Product", "Units", "Unit Price", "Total"])
    for row, (product, units, price) in enumerate(
        [
            ("Laptop Stand", 12, 24.5),
            ("USB-C Hub", 30, 18.0),
            ("Monitor Arm", 7, 89.99),
            ("Keyboard", 40, 55.0),
        ],
        start=2,
    ):
        sheet.append([product, units, price, f"=B{row}*C{row}"])

    # The row someone typed a number into instead of leaving the formula.
    sheet.append(["Webcam", 18, 42.0, 756.0])

    return _save(book, directory, "formulas_last_row_overwritten.xlsx")


def blank_rows_inside(directory: Path) -> Path:
    """Data split by a blank row in the middle.

    A gap must not be read as the end of the sheet: the last row of data is
    row 8, not row 4. Header row 1, data in rows 2 to 4 and 6 to 8.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Sales"

    sheet.append(["ID", "Product", "Region"])
    sheet.append([1001, "Laptop Stand", "EU"])
    sheet.append([1002, "USB-C Hub", "EU"])
    sheet.append([1003, "Monitor Arm", "US"])
    sheet.append([])
    sheet.append([1004, "Keyboard", "APAC"])
    sheet.append([1005, "Webcam", "US"])
    sheet.append([1006, "Headset", "EU"])

    return _save(book, directory, "blank_rows_inside.xlsx")


def multi_sheet(directory: Path) -> Path:
    """Two sheets with different columns.

    The tools work on the active sheet and never name one, so only Sales is
    ever reached. Notes is here so a test can show that the second sheet is
    invisible.
    """
    book = Workbook()

    sales = book.active
    sales.title = "Sales"
    sales.append(["ID", "Product", "Region"])
    sales.append([1001, "Laptop Stand", "EU"])
    sales.append([1002, "USB-C Hub", "EU"])

    notes = book.create_sheet("Notes")
    notes.append(["Author", "Comment"])
    notes.append(["Joori", "second sheet, not the active one"])

    return _save(book, directory, "multi_sheet.xlsx")


def formatting_past_data(directory: Path) -> Path:
    """Five rows of data in a sheet styled down to row 500.

    A filled cell far below the table makes max_row report 500, so anything
    trusting max_row would place a new row at 501 and leave a few hundred
    blank rows above it. The last row of data is 6.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Sales"

    sheet.append(["ID", "Product", "Region", "Order Date"])
    sheet.append([1001, "Laptop Stand", "EU", datetime(2026, 1, 14)])
    sheet.append([1002, "USB-C Hub", "EU", datetime(2026, 1, 16)])
    sheet.append([1003, "Monitor Arm", "US", datetime(2026, 1, 19)])
    sheet.append([1004, "Keyboard", "APAC", datetime(2026, 2, 5)])
    sheet.append([1005, "Webcam", "US", datetime(2026, 3, 3)])

    # Colour with no value in it, which is what a styled empty sheet leaves
    # behind and what makes max_row untrustworthy.
    sheet["A500"].fill = PatternFill("solid", fgColor="FFFF00")

    return _save(book, directory, "formatting_past_data.xlsx")