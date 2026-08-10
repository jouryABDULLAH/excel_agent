"""Tests for the Excel access layer.

These call workbook.py directly and hand it a path, so nothing here depends on
which file the tools reach for by default.

Several tests below assert on a wrong answer. Where that happens the test says
so in its name and in a comment: the sheet is misread currently, and the test is
there to make the misreading visible. will be fixed later.
"""

import hashlib

import make_fixtures
import pytest
from openpyxl import Workbook, load_workbook

from excel_agent import workbook as workbook_module
from excel_agent.workbook import (
    DEFAULT_HEADER_ROW,
    backup_once,
    backups_of,
    find_header_row,
    formula_columns,
    header_map,
    resolve_sheet,
    is_blank,
    last_data_row,
    load_book,
    load_values,
    save,
)


def digest(path) -> str:
    """The contents of a file as one comparable string."""
    return hashlib.sha256(path.read_bytes()).hexdigest()


def active_sheet(path):
    """The sheet the tools would work on, opened with formulas kept."""
    return load_book(path).active


# Finding the header row


def test_header_in_the_first_row_is_found(tmp_path):
    sheet = active_sheet(make_fixtures.clean_table(tmp_path))

    assert find_header_row(sheet) == 1
    assert list(header_map(sheet, 1)) == [
        "ID",
        "Product",
        "Region",
        "Units",
        "Unit Price",
    ]


def test_a_title_above_the_header_is_stepped_over(tmp_path):
    sheet = active_sheet(make_fixtures.title_above_header(tmp_path))

    # A title fills one cell, and a header needs two, so row 1 is passed over
    # and the blank row 2 with it.
    assert find_header_row(sheet) == 3
    assert list(header_map(sheet, 3)) == ["ID", "Product", "Region", "Units"]


def test_a_numeric_column_name_defeats_detection(tmp_path):
    sheet = active_sheet(make_fixtures.numeric_headers(tmp_path))

    # The real header is row 2, but 2024 is a number and a header is required
    # to be text the whole way across, so row 2 is skipped and the fallback
    # lands on the title. Wrong answer, pinned here on purpose.
    assert find_header_row(sheet) == DEFAULT_HEADER_ROW
    assert header_map(sheet, DEFAULT_HEADER_ROW) == {"Revenue by year": 1}


def test_a_single_column_defeats_detection(tmp_path):
    sheet = active_sheet(make_fixtures.one_column(tmp_path))

    # One column means one filled cell per row, and a header is required to
    # have two, so no row here can ever be taken for a header. The fallback
    # then reads the title as the column name. Wrong answer, pinned on purpose.
    assert find_header_row(sheet) == DEFAULT_HEADER_ROW
    assert header_map(sheet, DEFAULT_HEADER_ROW) == {"Inventory": 1}


def test_a_sheet_with_no_header_reads_its_first_row_as_one(tmp_path):
    sheet = active_sheet(make_fixtures.headerless_grid(tmp_path))

    # Nothing names a column, so the fallback turns the first row of data into
    # the column names and that row stops being data. Silent, which is why the
    # roadmap treats headerless sheets as unsupported until the grid tools
    # arrive and this fallback is replaced.
    assert find_header_row(sheet) == DEFAULT_HEADER_ROW
    assert list(header_map(sheet, DEFAULT_HEADER_ROW)) == [
        "Laptop Stand",
        "12",
        "24.5",
    ]


# Picking a sheet


def test_the_workbook_opens_on_its_first_sheet_when_none_is_named(tmp_path):
    book = load_book(make_fixtures.multi_sheet(tmp_path))

    assert resolve_sheet(book).title == "Sales"
    assert resolve_sheet(book, None).title == "Sales"
    assert resolve_sheet(book, "  ").title == "Sales"


def test_a_sheet_can_be_picked_by_name(tmp_path):
    book = load_book(make_fixtures.multi_sheet(tmp_path))

    assert resolve_sheet(book, "Notes").title == "Notes"
    # Matched the way a workbook name is, so the same spelling works in both.
    assert resolve_sheet(book, "notes").title == "Notes"
    assert resolve_sheet(book, "  NOTES  ").title == "Notes"


def test_a_name_reaching_no_sheet_says_which_ones_exist(tmp_path):
    book = load_book(make_fixtures.multi_sheet(tmp_path))

    with pytest.raises(ValueError) as refused:
        resolve_sheet(book, "Summary")

    assert 'There is no sheet called "Summary"' in str(refused.value)
    assert "Sales, Notes" in str(refused.value)


# Cells that look empty
#
# These sheets are built in memory and never saved, because openpyxl drops an
# empty string on the way out to a file: write "" and read it back and it is
# None again. An empty string only ever arrives from a file written by Excel
# or by another program, which is the case being pinned down here.


def sheet_of_rows(rows):
    """An unsaved sheet holding the given rows, exactly as written."""
    book = Workbook()
    sheet = book.active
    for row in rows:
        sheet.append(row)
    return sheet


def test_nothing_at_all_and_nothing_worth_reading_are_the_same_thing():
    assert is_blank(None)
    assert is_blank("")
    assert is_blank("   ")

    # A zero is a value someone meant, and so is False.
    assert not is_blank(0)
    assert not is_blank(False)
    assert not is_blank("EU")


def test_a_row_of_empty_strings_is_not_taken_for_the_header():
    sheet = sheet_of_rows(
        [
            ["", "  "],
            ["ID", "Product"],
            [1001, "Laptop Stand"],
        ]
    )

    # Two filled cells, both text, something underneath it: a row of empty
    # strings passes every test a header row is put through, and would be
    # adopted as one if blank meant only None.
    assert find_header_row(sheet) == 2
    assert header_map(sheet, 2) == {"ID": 1, "Product": 2}


def test_an_empty_string_does_not_name_a_column():
    sheet = sheet_of_rows(
        [
            ["ID", "", "Region"],
            [1001, "ignored", "EU"],
        ]
    )

    # A column called "" is one nothing could ask for by name, and it would
    # sit in the middle of every table read from the sheet.
    assert header_map(sheet, 1) == {"ID": 1, "Region": 3}


def test_a_row_cleared_to_empty_strings_is_not_the_end_of_the_data():
    sheet = sheet_of_rows(
        [
            ["ID", "Product"],
            [1001, "Laptop Stand"],
            [1002, "USB-C Hub"],
            ["", ""],
            ["  ", ""],
        ]
    )

    # Clearing a row and deleting it are different things to do, but afterwards
    # the sheet cannot tell them apart and neither can the model: the table
    # shows an empty row either way. So a row cleared to empty strings is
    # passed over exactly like one cleared to nothing, and the next row added
    # takes its place rather than landing below it.
    assert last_data_row(sheet, header_row=1) == 3


# Finding the last row of data


def test_a_blank_row_is_not_the_end_of_the_data(tmp_path):
    sheet = active_sheet(make_fixtures.blank_rows_inside(tmp_path))

    assert last_data_row(sheet, header_row=1) == 8


def test_formatting_below_the_data_is_not_counted_as_data(tmp_path):
    sheet = active_sheet(make_fixtures.formatting_past_data(tmp_path))

    # A coloured cell in row 500 is enough for openpyxl to count 500 rows.
    assert sheet.max_row == 500
    assert last_data_row(sheet, header_row=1) == 6


def test_a_sheet_with_no_data_returns_its_header_row(tmp_path):
    path = make_fixtures.clean_table(tmp_path)
    book = load_book(path)
    book.active.delete_rows(2, 5)
    save(book, path)

    assert last_data_row(active_sheet(path), header_row=1) == 1


# Spotting calculated columns


def test_a_column_of_formulas_is_reported_as_calculated(tmp_path):
    sheet = active_sheet(make_fixtures.formulas_last_row_overwritten(tmp_path))

    # Row 5 still holds its formula, so reading the column from there sees
    # Total, column 4, for what it is.
    assert formula_columns(sheet, header_row=1, last_row=5) == {4}


def test_a_formula_column_typed_over_in_its_last_row_is_missed(tmp_path):
    sheet = active_sheet(make_fixtures.formulas_last_row_overwritten(tmp_path))

    # Row 6 was typed over by hand, and the last row is the only row this
    # looks at, so the calculated column disappears from the answer. The
    # formulas in rows 2 to 5 are left unguarded by it, which is the reason
    # edit checks the target cell itself as well.
    assert formula_columns(sheet, header_row=1, last_row=6) == set()


def test_a_sheet_with_no_data_has_no_calculated_columns(tmp_path):
    sheet = active_sheet(make_fixtures.clean_table(tmp_path))

    assert formula_columns(sheet, header_row=1, last_row=1) == set()


# Backing up


def test_the_first_backup_copies_the_file_into_the_backups_folder(tmp_path):
    path = make_fixtures.clean_table(tmp_path)
    before = digest(path)

    backup = backup_once(path)

    assert backup.parent == workbook_module.BACKUP_DIR
    # Named for the workbook and the moment, and still a workbook, so it can
    # be opened by double clicking it.
    assert backup.name.startswith("clean_table-")
    assert backup.suffix == ".xlsx"
    assert digest(backup) == before
    # The workbook's own folder is left alone, which is what stops a backup
    # being mistaken for something to work on.
    assert list(tmp_path.glob("*.bak")) == []


def test_a_second_backup_of_the_same_workbook_is_not_taken(tmp_path):
    path = make_fixtures.clean_table(tmp_path)
    backup = backup_once(path)

    # The point of backing up once is that the copy is of the file as it was
    # found, so a later write cannot overwrite it with a half edited version.
    first = digest(backup)
    book = load_book(path)
    book.active["B2"] = "changed since the backup"
    save(book, path)

    assert backup_once(path) is None
    assert digest(backup) == first
    assert backups_of(path) == [backup]


def test_each_workbook_is_backed_up_in_its_own_right(tmp_path):
    sales = make_fixtures.clean_table(tmp_path)
    stock = make_fixtures.one_column(tmp_path)

    # Backing up once is per workbook, not per session, or the second file
    # written to in a session would never be copied at all.
    assert backup_once(sales) is not None
    assert backup_once(stock) is not None

    assert len(backups_of(sales)) == 1
    assert len(backups_of(stock)) == 1


def test_only_the_most_recent_backups_are_kept(tmp_path):
    path = make_fixtures.clean_table(tmp_path)

    # Stood in for by hand, because backups are stamped to the second and a
    # test cannot wait around for the clock to move on.
    workbook_module.BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    older = []
    for moment in range(1, 6):
        stale = workbook_module.BACKUP_DIR / f"clean_table-20200101-00000{moment}.xlsx"
        stale.write_bytes(b"an older backup")
        older.append(stale)

    kept = backup_once(path)

    assert len(backups_of(path)) == 3
    # The newest survive, and the one just taken is the newest of all.
    assert backups_of(path) == [older[3], older[4], kept]
    assert not older[0].exists()


def test_backups_of_one_workbook_are_not_counted_as_another_s(tmp_path):
    path = make_fixtures.clean_table(tmp_path)

    workbook_module.BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    # A workbook can be named after a backup, stamp and all. Its backups carry
    # a second stamp, and matching on the stamp exactly is what keeps the two
    # sets apart when the older ones are cleared out.
    impostor = workbook_module.BACKUP_DIR / "clean_table-20200101-000001-20200202-000002.xlsx"
    impostor.write_bytes(b"a backup of a differently named workbook")

    backup_once(path)

    assert impostor not in backups_of(path)
    assert impostor.exists()


# Saving


def test_saving_writes_the_change_and_leaves_no_temporary_file(tmp_path):
    path = make_fixtures.clean_table(tmp_path)

    book = load_book(path)
    book.active["B2"] = "Standing Desk"
    save(book, path)

    assert load_workbook(path).active["B2"].value == "Standing Desk"
    assert list(tmp_path.glob("*.writing")) == []


def test_a_save_that_fails_leaves_the_original_file_alone(tmp_path):
    path = make_fixtures.clean_table(tmp_path)
    before = digest(path)

    book = load_book(path)
    book.active["B2"] = "never written"

    def fails(*arguments, **keywords):
        raise OSError("disk full")

    # Writing to one side and renaming means a failed write cannot touch the
    # real file, only the temporary one. The temporary file is left behind,
    # which is untidy but harmless, and is not asserted on here.
    book.save = fails
    with pytest.raises(OSError):
        save(book, path)

    assert digest(path) == before


def test_saving_a_workbook_opened_for_reading_is_refused(tmp_path):
    path = make_fixtures.formulas_last_row_overwritten(tmp_path)
    before = digest(path)

    # This workbook holds the results of the formulas rather than the formulas
    # themselves, so saving it would replace every formula with the number it
    # last worked out.
    book = load_values(path)

    with pytest.raises(ValueError, match="load_book"):
        save(book, path)

    assert digest(path) == before
    assert list(tmp_path.glob("*.writing")) == []
