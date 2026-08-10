"""Tests for the chart tool.

Charts are checked after the file has been saved and read back, not on the
workbook still in memory, because what matters is that the chart is in the
file the user opens.
"""

import hashlib

import make_fixtures
from openpyxl import load_workbook

from excel_agent.tools.charts import apply_chart_change, modify_chart


def digest(path) -> str:
    """The contents of a file as one comparable string."""
    return hashlib.sha256(path.read_bytes()).hexdigest()


def charts_in(path, sheet=None) -> list:
    """The charts in the saved file."""
    book = load_workbook(path)
    worksheet = book[sheet] if sheet else book.active
    return worksheet._charts


def title_of(chart) -> str:
    """The words openpyxl buries several layers down inside a chart title."""
    return chart.title.tx.rich.p[0].r[0].t


# Drawing


def test_a_chart_is_drawn_and_stays_in_the_file(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    answer = modify_chart.invoke(
        {"action": "add", "values": "Units", "categories": "Product"}
    )

    assert "Drew a bar chart of Units, labelled by Product" in answer
    assert "(Sales in clean_table.xlsx)" in answer
    assert "covers rows 2 to 6" in answer

    charts = charts_in(path)
    assert len(charts) == 1
    assert title_of(charts[0]) == "Units by Product"


def test_a_chart_is_placed_clear_of_the_data(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    # Five columns of data, so the chart starts at G and never covers them.
    answer = modify_chart.invoke({"action": "add", "values": "Units"})

    assert "at G1" in answer
    assert charts_in(path)[0].anchor._from.col == 6


def test_each_kind_of_chart_can_be_drawn(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    for kind, expected in (("bar", "barChart"), ("line", "lineChart"), ("pie", "pieChart")):
        modify_chart.invoke({"action": "remove"})
        modify_chart.invoke({"action": "add", "values": "Units", "kind": kind})
        assert charts_in(path)[0].tagname == expected


def test_a_chart_can_be_given_its_own_title(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    modify_chart.invoke(
        {"action": "add", "values": "Units", "title": "How many we sold"}
    )

    assert title_of(charts_in(path)[0]) == "How many we sold"


def test_a_calculated_column_can_be_charted(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.formulas_all_the_way_down(tmp_path))

    # Nothing can be read out of Total, because its results are not stored in
    # the file. A chart points at the cells rather than their values, so Excel
    # works it out when the file is opened.
    answer = modify_chart.invoke(
        {"action": "add", "values": "Total", "categories": "Product"}
    )

    assert "Drew a bar chart of Total" in answer
    assert len(charts_in(path)) == 1


def test_a_chart_survives_a_later_edit(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    modify_chart.invoke({"action": "add", "values": "Units"})

    from excel_agent.tools.modify import modify_sheet

    modify_sheet.invoke({"action": "edit", "row": 2, "values": {"Units": 99}})

    # Every tool call loads the file and saves it again. A chart that did not
    # survive that would vanish the moment anything else was changed.
    assert len(charts_in(path)) == 1
    assert load_workbook(path).active["D2"].value == 99


# Where a chart goes


def test_a_second_chart_goes_below_the_first(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    first = modify_chart.invoke({"action": "add", "values": "Units", "kind": "bar"})
    second = modify_chart.invoke(
        {"action": "add", "values": "Unit Price", "kind": "pie"}
    )

    # Both are in the file, and the second is a chart's depth further down, so
    # it does not sit on top of the first and hide it.
    assert "at G1" in first
    assert "at G17" in second

    charts = charts_in(path)
    assert [chart.tagname for chart in charts] == ["barChart", "pieChart"]
    assert [chart.anchor._from.row for chart in charts] == [0, 16]


def test_charts_drawn_in_a_later_session_still_step_down(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    modify_chart.invoke({"action": "add", "values": "Units"})

    # Each call opens the file again, so the chart already saved has to be
    # counted from the file rather than from anything held in memory.
    answer = modify_chart.invoke({"action": "add", "values": "Unit Price"})

    assert "at G17" in answer
    assert len(charts_in(path)) == 2


def test_a_chart_can_be_put_where_it_is_asked_for(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    answer = modify_chart.invoke(
        {"action": "add", "values": "Units", "anchor": "b20"}
    )

    assert "at B20" in answer
    placed = charts_in(path)[0].anchor._from
    assert (placed.col, placed.row) == (1, 19)


def test_something_that_is_not_a_cell_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_chart.invoke(
        {"action": "add", "values": "Units", "anchor": "over there"}
    )

    assert "is not a cell" in answer
    assert digest(path) == before


# Taking charts away


def test_charts_can_be_taken_off_a_sheet(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    modify_chart.invoke({"action": "add", "values": "Units"})

    answer = modify_chart.invoke({"action": "remove"})

    assert "Removed 1 chart" in answer
    assert charts_in(path) == []
    # The data is left exactly where it was.
    assert load_workbook(path).active["D2"].value == 12


def test_removing_charts_when_there_are_none_says_so(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_chart.invoke({"action": "remove"})

    assert "no charts" in answer
    assert digest(path) == before


# Refusing


def test_a_column_of_text_cannot_be_plotted(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_chart.invoke({"action": "add", "values": "Product"})

    assert '"Product" cannot be plotted' in answer
    assert digest(path) == before


def test_an_unknown_column_is_named_and_the_real_ones_listed(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_chart.invoke({"action": "add", "values": "Profit"})

    assert "Unknown column(s): Profit" in answer
    assert digest(path) == before


def test_an_unknown_kind_of_chart_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    # Called underneath the tool, because the tool's own argument checking
    # rejects anything but bar, line and pie before the function runs. This is
    # the last line of defence, for a caller that is not the model.
    answer = apply_chart_change("add", "Units", None, "scatter", None, None, path)

    assert "bar, line, pie" in answer
    assert digest(path) == before


def test_an_action_that_is_neither_add_nor_remove_changes_nothing(
    tmp_path, use_workbook
):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = apply_chart_change("redraw", "Units", None, "bar", None, None, path)

    assert 'Unknown action "redraw"' in answer
    assert digest(path) == before


def test_drawing_needs_a_column_to_plot(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_chart.invoke({"action": "add"})

    assert "needs the name of a column to plot" in answer
    assert digest(path) == before


# Choosing a sheet


def test_a_named_sheet_is_the_one_charted(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.multi_sheet(tmp_path))

    answer = modify_chart.invoke(
        {"action": "add", "values": "ID", "sheet": "Sales"}
    )

    assert "(Sales in multi_sheet.xlsx)" in answer
    assert len(charts_in(path, "Sales")) == 1
    assert charts_in(path, "Notes") == []


def test_a_sheet_that_does_not_exist_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.multi_sheet(tmp_path))
    before = digest(path)

    answer = modify_chart.invoke(
        {"action": "add", "values": "ID", "sheet": "Summary"}
    )

    assert 'There is no sheet called "Summary"' in answer
    assert digest(path) == before