"""Tests for the two tools the agent is given.

Everything here goes through .invoke, the way the model calls them, so what is
asserted on is the string the model would actually read back.

The tools take no path and reach the workbook through workbook.py, so every
test points them at a sheet of its own first with use_workbook. A test that
forgot to would be caught by the guard in conftest.py rather than by the tools.
"""

import hashlib

import make_fixtures
from openpyxl import load_workbook

from excel_agent.tools.inspect import inspect_sheet
from excel_agent.tools.modify import apply_change, modify_sheet


def digest(path) -> str:
    """The contents of a file as one comparable string."""
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sheet_of(path):
    """The saved sheet, with formulas kept as formulas."""
    return load_workbook(path).active


# Reading


def test_the_whole_sheet_is_read_with_its_real_row_numbers(tmp_path, use_workbook):
    use_workbook(make_fixtures.clean_table(tmp_path))

    answer = inspect_sheet.invoke({})

    # The workbook is named alongside the sheet, so two tables read in one
    # conversation cannot be mistaken for one another.
    assert (
        "Sheet: Sales in clean_table.xlsx (5 rows of data, column names in row 1)"
        in answer
    )
    assert "| row | ID | Product | Region | Units | Unit Price |" in answer
    # Row 2 in the table is row 2 in Excel, which is what modify_sheet needs.
    assert "| 2 | 1001 | Laptop Stand | EU | 12 | 24.5 |" in answer
    assert "| 6 | 1005 | Webcam | US | 18 | 42 |" in answer


def test_columns_come_back_in_the_order_they_were_asked_for(tmp_path, use_workbook):
    use_workbook(make_fixtures.clean_table(tmp_path))

    answer = inspect_sheet.invoke({"columns": ["Region", "Product"]})

    assert "| row | Region | Product |" in answer
    assert "| 2 | EU | Laptop Stand |" in answer


def test_an_unknown_column_is_named_and_the_real_ones_listed(tmp_path, use_workbook):
    use_workbook(make_fixtures.clean_table(tmp_path))

    answer = inspect_sheet.invoke({"columns": ["Profit"]})

    assert "Unknown column(s): Profit" in answer
    assert "The sheet has: ID, Product, Region, Units, Unit Price." in answer


def test_asking_for_no_rows_explains_itself(tmp_path, use_workbook):
    use_workbook(make_fixtures.clean_table(tmp_path))

    assert "no rows were read" in inspect_sheet.invoke({"max_rows": 0})


def test_a_shortened_read_says_how_to_see_the_rest(tmp_path, use_workbook):
    use_workbook(make_fixtures.clean_table(tmp_path))

    answer = inspect_sheet.invoke({"max_rows": 2})

    assert "Showing rows 2 to 3." in answer
    assert "Rows 4 to 6 were not shown" in answer
    assert "start_row=4" in answer


def test_reading_past_the_end_says_where_the_data_ends(tmp_path, use_workbook):
    use_workbook(make_fixtures.clean_table(tmp_path))

    answer = inspect_sheet.invoke({"start_row": 500})

    assert "ending at row 6" in answer
    assert "nothing to read from row 500" in answer


def test_a_calculated_cell_is_shown_as_its_formula_with_a_warning(
    tmp_path, use_workbook
):
    use_workbook(make_fixtures.formulas_all_the_way_down(tmp_path))

    answer = inspect_sheet.invoke({})

    # Nothing has opened this file in Excel, so no result has ever been worked
    # out and cached. The formula itself is shown instead of a blank, with a
    # line telling the model not to try to set the cell.
    assert "=B2*C2" in answer
    assert "calculated by the sheet itself" in answer


def test_the_sheet_the_workbook_opens_on_is_read_when_none_is_named(
    tmp_path, use_workbook
):
    use_workbook(make_fixtures.multi_sheet(tmp_path))

    answer = inspect_sheet.invoke({})

    # Leaving the sheet out reads the one the file opens on, so a single sheet
    # workbook behaves exactly as it did before sheets could be named.
    assert "Sheet: Sales" in answer
    assert "Author" not in answer


# Adding


def test_a_new_row_goes_under_the_last_one(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    answer = modify_sheet.invoke(
        {"action": "add", "values": {"Product": "Standing Desk", "Region": "EU"}}
    )

    assert answer.startswith("Added row 7 with Product = Standing Desk, Region = EU.")
    assert "Any other column was left blank." in answer

    sheet = sheet_of(path)
    assert sheet["B7"].value == "Standing Desk"
    assert sheet["A7"].value is None


def test_a_new_row_picks_up_the_calculated_column(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.formulas_all_the_way_down(tmp_path))

    answer = modify_sheet.invoke(
        {"action": "add", "values": {"Product": "Standing Desk", "Units": 4, "Unit Price": 120.0}}
    )

    assert "Copied the formula in Total down from row 6." in answer
    # The references shift with the row, so row 7 multiplies row 7.
    assert sheet_of(path)["D7"].value == "=B7*C7"


def test_a_calculated_column_typed_over_in_its_last_row_is_not_copied_down(
    tmp_path, use_workbook
):
    path = use_workbook(make_fixtures.formulas_last_row_overwritten(tmp_path))

    answer = modify_sheet.invoke({"action": "add", "values": {"Product": "Standing Desk"}})

    # The formula is copied from the last row, and the last row holds a number
    # someone typed, so the new row gets a blank where the sheet's other rows
    # have a formula. Recorded here as it stands today.
    assert "Copied the formula" not in answer
    assert sheet_of(path)["D7"].value is None


def test_a_new_row_ignores_formatting_far_below_the_data(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.formatting_past_data(tmp_path))

    answer = modify_sheet.invoke({"action": "add", "values": {"Product": "Standing Desk"}})

    # Row 500 is coloured in, so trusting max_row would put this at row 501
    # and leave several hundred blank rows above it.
    assert "Added row 7" in answer
    assert sheet_of(path)["B7"].value == "Standing Desk"


def test_setting_a_calculated_column_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.formulas_all_the_way_down(tmp_path))
    before = digest(path)

    answer = modify_sheet.invoke(
        {"action": "add", "values": {"Product": "Standing Desk", "Total": 480}}
    )

    assert "Total is worked out by a formula" in answer
    assert digest(path) == before


def test_adding_without_values_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_sheet.invoke({"action": "add", "values": {}})

    assert "needs at least one column in values" in answer
    assert digest(path) == before


def test_adding_an_unknown_column_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_sheet.invoke({"action": "add", "values": {"Profit": 10}})

    assert "Unknown column(s): Profit" in answer
    assert digest(path) == before


# Editing


def test_editing_changes_the_columns_given_and_no_others(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    answer = modify_sheet.invoke({"action": "edit", "row": 3, "values": {"Units": 99}})

    assert answer == "Updated row 3: Units = 99. (Sales in clean_table.xlsx)"

    sheet = sheet_of(path)
    assert sheet["D3"].value == 99
    assert sheet["B3"].value == "USB-C Hub"


def test_a_cell_can_be_cleared(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    answer = modify_sheet.invoke({"action": "edit", "row": 3, "values": {"Region": None}})

    assert answer == "Updated row 3: Region = (blank). (Sales in clean_table.xlsx)"
    assert sheet_of(path)["C3"].value is None


def test_editing_a_cell_in_a_calculated_column_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.formulas_all_the_way_down(tmp_path))
    before = digest(path)

    answer = modify_sheet.invoke({"action": "edit", "row": 3, "values": {"Total": 500}})

    assert "Total is worked out by a formula" in answer
    assert digest(path) == before


def test_a_formula_partway_down_a_column_is_protected(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.formulas_last_row_overwritten(tmp_path))
    before = digest(path)

    # Row 3 holds a formula, but the last row of Total holds a number someone
    # typed, so the column does not look calculated and the column level check
    # lets this through. Looking at the cell itself is what catches it.
    answer = modify_sheet.invoke(
        {"action": "edit", "row": 3, "values": {"Units": 99, "Total": 500}}
    )

    assert "Total is worked out by a formula" in answer
    # Units was in the same call and is an ordinary column, so its absence from
    # the file is what shows the refusal happened before anything was written.
    assert digest(path) == before


def test_a_number_that_replaced_a_formula_can_still_be_corrected(
    tmp_path, use_workbook
):
    path = use_workbook(make_fixtures.formulas_last_row_overwritten(tmp_path))

    # Row 6's Total is a number, not a formula, so there is no calculation to
    # protect and correcting it is allowed. Checking the cell rather than the
    # column is what makes the difference between this and the test above.
    answer = modify_sheet.invoke({"action": "edit", "row": 6, "values": {"Total": 800}})

    assert answer == (
        "Updated row 6: Total = 800. "
        "(Sales in formulas_last_row_overwritten.xlsx)"
    )
    assert sheet_of(path)["D6"].value == 800


def test_editing_a_row_that_does_not_exist_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_sheet.invoke({"action": "edit", "row": 9999, "values": {"Units": 1}})

    assert "Row 9999 does not exist. The sheet has rows 2 to 6." in answer
    assert digest(path) == before


def test_editing_the_header_row_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    # Row 1 holds the column names, and renaming a column is not something
    # these tools do.
    answer = modify_sheet.invoke({"action": "edit", "row": 1, "values": {"Units": 1}})

    assert "Row 1 does not exist" in answer
    assert digest(path) == before


def test_editing_without_a_row_number_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_sheet.invoke({"action": "edit", "values": {"Units": 1}})

    assert "needs a row number" in answer
    assert digest(path) == before


# Removing


def test_removing_a_row_shifts_the_rest_up_and_says_so(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    answer = modify_sheet.invoke({"action": "remove", "row": 3})

    assert answer.startswith("Removed row 3.")
    assert "now out of date" in answer

    sheet = sheet_of(path)
    # What was row 4 is row 3 now, which is exactly why the message warns.
    assert sheet["B3"].value == "Monitor Arm"
    assert sheet.max_row == 5


def test_removing_a_row_that_does_not_exist_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_sheet.invoke({"action": "remove", "row": 9999})

    assert "does not exist" in answer
    assert digest(path) == before


def test_removing_without_a_row_number_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_sheet.invoke({"action": "remove"})

    assert "needs a row number" in answer
    assert digest(path) == before


def test_an_action_that_is_not_one_of_the_three_changes_nothing(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    # Called underneath the tool on purpose: the tool's own argument checking
    # rejects anything but add, edit and remove before the function runs, so
    # this last line of defence cannot be reached through .invoke.
    answer = apply_change("sort", row=2, values=None, path=path)

    assert 'Unknown action "sort"' in answer
    assert digest(path) == before


# Choosing a workbook


def test_a_workbook_can_be_named(tmp_path, use_workbook):
    use_workbook(make_fixtures.clean_table(tmp_path))
    make_fixtures.blank_rows_inside(tmp_path)

    # Both files sit in the same folder, so naming one is the only way to read
    # anything but the default.
    answer = inspect_sheet.invoke({"workbook": "blank_rows_inside"})

    assert "in blank_rows_inside.xlsx" in answer
    assert "Headset" in answer


def test_a_named_workbook_is_the_one_that_gets_changed(tmp_path, use_workbook):
    default = use_workbook(make_fixtures.clean_table(tmp_path))
    named = make_fixtures.blank_rows_inside(tmp_path)
    untouched = digest(default)

    answer = modify_sheet.invoke(
        {"action": "edit", "row": 2, "values": {"Product": "Standing Desk"},
         "workbook": "blank_rows_inside.xlsx"}
    )

    assert "Updated row 2" in answer
    assert sheet_of(named)["B2"].value == "Standing Desk"
    # The default workbook was not the one named, so it was not written to.
    assert digest(default) == untouched


def test_leaving_the_workbook_out_uses_the_one_being_worked_on(tmp_path, use_workbook):
    default = use_workbook(make_fixtures.clean_table(tmp_path))
    other = make_fixtures.blank_rows_inside(tmp_path)
    untouched = digest(other)

    modify_sheet.invoke({"action": "edit", "row": 2, "values": {"Product": "Standing Desk"}})

    assert sheet_of(default)["B2"].value == "Standing Desk"
    assert digest(other) == untouched


def test_a_workbook_that_does_not_exist_is_named_along_with_the_ones_that_do(
    tmp_path, use_workbook
):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_sheet.invoke(
        {"action": "remove", "row": 2, "workbook": "invoices"}
    )

    assert 'There is no workbook called "invoices"' in answer
    assert "clean_table.xlsx" in answer
    assert digest(path) == before


def test_a_name_reaching_outside_the_data_folder_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    # A name arrives from the model, so one carrying a folder of its own must
    # not be followed wherever it points.
    answer = modify_sheet.invoke(
        {"action": "remove", "row": 2, "workbook": "../../secrets.xlsx"}
    )

    assert "is not a workbook name" in answer
    assert digest(path) == before


# Choosing a sheet


def test_a_named_sheet_is_read(tmp_path, use_workbook):
    use_workbook(make_fixtures.multi_sheet(tmp_path))

    answer = inspect_sheet.invoke({"sheet": "Notes"})

    assert "Sheet: Notes in multi_sheet.xlsx" in answer
    assert "| row | Author | Comment |" in answer
    assert "Joori" in answer


def test_a_sheet_name_is_matched_however_it_is_spelled(tmp_path, use_workbook):
    use_workbook(make_fixtures.multi_sheet(tmp_path))

    assert "Sheet: Notes" in inspect_sheet.invoke({"sheet": "notes"})
    assert "Sheet: Notes" in inspect_sheet.invoke({"sheet": "  NOTES  "})


def test_a_named_sheet_is_the_one_that_gets_changed(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.multi_sheet(tmp_path))

    answer = modify_sheet.invoke(
        {"action": "edit", "row": 2, "values": {"Comment": "changed"}, "sheet": "Notes"}
    )

    assert "Updated row 2" in answer

    book = load_workbook(path)
    assert book["Notes"]["B2"].value == "changed"
    # Both sheets have a row 2, and the one that was not named still holds
    # what it always held.
    assert book["Sales"]["B2"].value == "Laptop Stand"


def test_a_row_added_to_a_named_sheet_lands_in_that_sheet(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.multi_sheet(tmp_path))

    # Notes ends at row 2 and Sales at row 3, so the row number in the answer
    # says which sheet was measured for the end of the data.
    answer = modify_sheet.invoke(
        {"action": "add", "values": {"Author": "Sam"}, "sheet": "Notes"}
    )

    assert "Added row 3" in answer

    book = load_workbook(path)
    assert book["Notes"]["A3"].value == "Sam"
    assert book["Sales"].max_row == 3


def test_a_sheet_that_does_not_exist_is_refused_by_both_tools(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.multi_sheet(tmp_path))
    before = digest(path)

    read = inspect_sheet.invoke({"sheet": "Summary"})
    written = modify_sheet.invoke({"action": "remove", "row": 2, "sheet": "Summary"})

    for answer in (read, written):
        assert 'There is no sheet called "Summary"' in answer
        assert "Sales, Notes" in answer

    assert digest(path) == before
