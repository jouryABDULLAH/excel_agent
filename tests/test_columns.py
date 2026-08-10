"""Tests for the column tool.

Everything goes through .invoke, the way the model calls it, so what is
asserted on is the string the model would read back. A refusal is checked
against the file as well as against the words: an explanation is only true if
nothing was written.
"""

import hashlib

import make_fixtures
from openpyxl import load_workbook

from excel_agent.tools.columns import apply_column_change, modify_column


def digest(path) -> str:
    """The contents of a file as one comparable string."""
    return hashlib.sha256(path.read_bytes()).hexdigest()


def headers_of(path, sheet=None) -> list:
    """The column names as they stand in the saved file."""
    book = load_workbook(path)
    worksheet = book[sheet] if sheet else book.active
    return [cell.value for cell in worksheet[1]]


# Adding


def test_a_new_column_goes_on_the_end(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    answer = modify_column.invoke({"action": "add", "name": "Notes"})

    assert 'Added a column called "Notes", at F.' in answer
    assert "use modify_sheet to put values into it" in answer
    assert headers_of(path) == ["ID", "Product", "Region", "Units", "Unit Price", "Notes"]


def test_a_new_column_arrives_empty(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    modify_column.invoke({"action": "add", "name": "Notes"})

    # Only the header cell is written. The rows below stay as they were.
    sheet = load_workbook(path).active
    assert [sheet.cell(row=row, column=6).value for row in range(2, 7)] == [None] * 5


def test_a_second_column_of_the_same_name_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_column.invoke({"action": "add", "name": "Region"})

    assert 'There is already a column called "Region"' in answer
    assert digest(path) == before


def test_a_column_needs_a_name(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    for arguments in ({"action": "add"}, {"action": "add", "name": "   "}):
        assert "needs the name of a column" in modify_column.invoke(arguments)

    assert digest(path) == before


# Renaming


def test_renaming_keeps_the_data_where_it_is(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    answer = modify_column.invoke(
        {"action": "rename", "name": "Units", "new_name": "Quantity"}
    )

    assert 'Renamed the column "Units" to "Quantity".' in answer
    assert headers_of(path) == [
        "ID",
        "Product",
        "Region",
        "Quantity",
        "Unit Price",
    ]
    # The column kept its place, so what was under Units is under Quantity.
    assert load_workbook(path).active["D2"].value == 12


def test_renaming_a_column_that_is_not_there_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_column.invoke(
        {"action": "rename", "name": "Profit", "new_name": "Margin"}
    )

    assert 'There is no column called "Profit"' in answer
    assert "The sheet has: ID, Product, Region, Units, Unit Price." in answer
    assert digest(path) == before


def test_renaming_onto_a_name_already_taken_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_column.invoke(
        {"action": "rename", "name": "Units", "new_name": "Region"}
    )

    assert 'There is already a column called "Region"' in answer
    assert digest(path) == before


def test_renaming_needs_a_new_name(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_column.invoke({"action": "rename", "name": "Units"})

    assert "needs a new_name" in answer
    assert digest(path) == before


# Deleting


def test_a_column_nothing_depends_on_can_be_deleted(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))

    answer = modify_column.invoke({"action": "delete", "name": "Region"})

    assert 'Deleted the column "Region"' in answer
    assert "cannot be brought back" in answer
    assert headers_of(path) == ["ID", "Product", "Units", "Unit Price"]
    # What was to the right of it has moved left, values and all.
    assert load_workbook(path).active["C2"].value == 12


def test_a_column_a_formula_reads_cannot_be_deleted(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.formulas_all_the_way_down(tmp_path))
    before = digest(path)

    answer = modify_column.invoke({"action": "delete", "name": "Units"})

    assert '"Units" cannot be deleted' in answer
    assert '"Total"' in answer
    assert "=B2*C2" in answer
    assert digest(path) == before


def test_a_column_no_formula_names_still_cannot_be_deleted_from_under_one(
    tmp_path, use_workbook
):
    path = use_workbook(make_fixtures.formulas_all_the_way_down(tmp_path))
    before = digest(path)

    # No formula mentions Product. Deleting it would still slide Units and
    # Unit Price one place left while the formulas that read them stayed put,
    # which is the failure openpyxl makes silently.
    answer = modify_column.invoke({"action": "delete", "name": "Product"})

    assert '"Product" cannot be deleted' in answer
    assert digest(path) == before


def test_a_calculated_column_can_be_deleted(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.formulas_all_the_way_down(tmp_path))

    # Total's own formulas go with it, and they only read columns to its left,
    # so nothing is left behind pointing anywhere wrong.
    answer = modify_column.invoke({"action": "delete", "name": "Total"})

    assert 'Deleted the column "Total"' in answer
    assert headers_of(path) == ["Product", "Units", "Unit Price"]


def test_a_formula_partway_down_a_column_still_blocks_a_deletion(
    tmp_path, use_workbook
):
    path = use_workbook(make_fixtures.formulas_last_row_overwritten(tmp_path))
    before = digest(path)

    # The last row of Total holds a number someone typed, so asking the column
    # whether it is calculated answers no. Every row is looked at instead.
    answer = modify_column.invoke({"action": "delete", "name": "Units"})

    assert '"Units" cannot be deleted' in answer
    assert digest(path) == before


def test_deleting_a_column_that_is_not_there_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    answer = modify_column.invoke({"action": "delete", "name": "Profit"})

    assert 'There is no column called "Profit"' in answer
    assert digest(path) == before


# Choosing a workbook and a sheet


def test_a_named_sheet_is_the_one_whose_columns_change(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.multi_sheet(tmp_path))

    answer = modify_column.invoke(
        {"action": "add", "name": "Checked", "sheet": "Notes"}
    )

    assert 'Added a column called "Checked", at C.' in answer
    assert headers_of(path, "Notes") == ["Author", "Comment", "Checked"]
    assert headers_of(path, "Sales") == ["ID", "Product", "Region"]


def test_a_sheet_that_does_not_exist_is_refused(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.multi_sheet(tmp_path))
    before = digest(path)

    answer = modify_column.invoke(
        {"action": "add", "name": "Checked", "sheet": "Summary"}
    )

    assert 'There is no sheet called "Summary"' in answer
    assert "Sales, Notes" in answer
    assert digest(path) == before


def test_an_action_that_is_not_one_of_the_three_changes_nothing(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    before = digest(path)

    # Called underneath the tool, because its own argument checking rejects
    # anything but add, rename and delete before the function runs.
    answer = apply_column_change("move", "Units", None, path)

    assert 'Unknown action "move"' in answer
    assert digest(path) == before