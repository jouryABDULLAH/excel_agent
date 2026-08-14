"""Tests for the conversation the agent keeps. Uses a fake model.
"""

import make_fixtures
from langchain.agents import create_agent
from langchain_core.language_models.chat_models import BaseChatModel
from langchain_core.messages import AIMessage, BaseMessage
from langchain_core.outputs import ChatGeneration, ChatResult
from langgraph.checkpoint.memory import InMemorySaver
from openpyxl import load_workbook

from scripted import ScriptedModel

from excel_agent.agent import answer_of, ask, new_thread, tool_calls_in
from excel_agent.tools import LOCAL_TOOLS


def agent_reading(script, tools=()):
    """An agent that will say the given things, keeping its conversations."""
    return create_agent(
        ScriptedModel(script=list(script)),
        list(tools),
        system_prompt="terse",
        checkpointer=InMemorySaver(),
    )


def calling_modify(call_id: str, **arguments) -> AIMessage:
    """A message asking for one modify_row call."""
    return AIMessage(
        content="",
        tool_calls=[{"name": "modify_row", "args": arguments, "id": call_id}],
    )


# Keeping a conversation


def test_a_thread_holds_on_to_what_was_said_before():
    agent = agent_reading([AIMessage("first"), AIMessage("second")])
    thread = new_thread()

    ask(agent, "my name is Joori", thread)
    ask(agent, "what is my name?", thread)

    said = agent.get_state({"configurable": {"thread_id": thread}}).values["messages"]
    assert [message.content for message in said] == [
        "my name is Joori",
        "first",
        "what is my name?",
        "second",
    ]


def test_a_new_thread_starts_with_nothing_said():
    agent = agent_reading([AIMessage("first"), AIMessage("second")])

    ask(agent, "my name is Joori", new_thread())
    forgotten = new_thread()
    ask(agent, "what is my name?", forgotten)

    said = agent.get_state({"configurable": {"thread_id": forgotten}}).values["messages"]
    assert [message.content for message in said] == ["what is my name?", "second"]


def test_only_this_turn_comes_back_from_ask():
    agent = agent_reading([AIMessage("first"), AIMessage("second")])
    thread = new_thread()

    ask(agent, "one", thread)
    produced = ask(agent, "two", thread)

    assert [message.content for message in produced] == ["second"]
    assert answer_of(produced) == "second"


# Tool calls through a kept conversation


def test_a_tool_call_and_its_answer_both_come_back(tmp_path, use_workbook):
    use_workbook(make_fixtures.clean_table(tmp_path))
    agent = agent_reading(
        [
            calling_modify("1", action="edit", row=2, values={"Units": 99}),
            AIMessage("Set the units on row 2 to 99."),
        ],
        tools=LOCAL_TOOLS,
    )

    produced = ask(agent, "set row 2 units to 99", new_thread())

    assert tool_calls_in(produced) == [
        "modify_row(action='edit', row=2, values={'Units': 99})"
    ]
    # The tool's own answer is in there too, between the call and the reply,
    # which is what the model reads before saying anything.
    assert any("Updated row 2" in str(message.content) for message in produced)
    assert answer_of(produced) == "Set the units on row 2 to 99."


def test_a_row_added_in_one_turn_can_be_removed_in_the_next(tmp_path, use_workbook):
    path = use_workbook(make_fixtures.clean_table(tmp_path))
    agent = agent_reading(
        [
            calling_modify(
                "1", action="add", values={"Product": "Monitor Arm", "Region": "US"}
            ),
            AIMessage("Added it as row 7."),
            calling_modify("2", action="remove", row=7),
            AIMessage("Removed row 7 again."),
        ],
        tools=LOCAL_TOOLS,
    )
    thread = new_thread()

    ask(agent, "add a row for a Monitor Arm in the US", thread)
    assert load_workbook(path).active["B7"].value == "Monitor Arm"

    ask(agent, "actually, remove that row again", thread)
    assert load_workbook(path).active["B7"].value is None


    said = agent.get_state({"configurable": {"thread_id": thread}}).values["messages"]
    assert len(said) == 8
    assert sum(1 for message in said if getattr(message, "tool_calls", None)) == 2